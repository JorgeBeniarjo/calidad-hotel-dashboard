import io

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Datos") -> bytes:
    """Convierte un DataFrame a bytes de Excel con cabecera formateada."""
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2A73CC")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            # Convertir NaN a None para celdas vacías
            if pd.isna(value) if not isinstance(value, str) else False:
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        series = df.iloc[:, col_idx - 1].astype(str)
        max_len = max(
            len(str(df.columns[col_idx - 1])),
            int(series.str.len().max()) if not df.empty else 0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    wb.save(output)
    return output.getvalue()


def download_excel_button(
    df: pd.DataFrame,
    filename: str,
    label: str = "⬇️ Descargar Excel",
    key: str = None,
    sheet_name: str = "Datos",
) -> None:
    """Renderiza un botón de descarga de Excel en Streamlit."""
    fname = filename if filename.endswith(".xlsx") else filename + ".xlsx"
    data = df_to_excel_bytes(df, sheet_name=sheet_name)
    st.download_button(
        label=label,
        data=data,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=key,
    )
